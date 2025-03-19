import openpyxl
import json
import re

class ProcessData:
    def __init__(self, docxl1, docxl2):
        # activar data de xl para su manipulación con openpyxl:
        self.docxl1, self.docxl2 = openpyxl.load_workbook(docxl1).active, openpyxl.load_workbook(docxl2).active
        
        # obtener dataframe xl en json y como una lista de diccionarios. Además definimos headers de cada dtfr:
        self.header1, self.jsondata1, self.dictdata1  = self.data_convert(self.docxl1)
        self.header2, self.jsondata2, self.dictdata2 = self.data_convert(self.docxl2)
        #self.test_var = self.data_convert(self.docxl2)

        # definir una lista con los bnc y bn para comprobar datos faltantes:
        # self.bnc1, self.bn1 = self.define_columns(self.dictdata1, 'booking_number_carrier', 'booking_number')
        # self.bnc2, self.bn2 = self.define_columns(self.dictdata2, 'C', 'AA')

        # definir una lista de diccionarios con bnc como key y otro dict como value con respectivo bnc y bn:
        self.cut_dictdata1 = self.cut_dict(self.dictdata1, 'booking_number_carrier', 'booking_number')
        self.cut_dictdata2 = self.cut_dict(self.dictdata2, 'C', 'AA')

    @staticmethod
    def data_convert(dataframe):
        datajsonlist = []
        datadictlist = []
        headers = [cell.value for cell in dataframe[1]]
        # header_key = header.index("Nº Boleto")                
        
        for row in range(1, dataframe.max_row):
            dict = {}
            for col, h in zip(dataframe.iter_cols(1, dataframe.max_column), headers):
                dict[h] = col[row].value
            datadictlist.append(dict)
            jsonelement = json.dumps(dict, indent=3, default=str)
            datajsonlist.append(jsonelement)

        return headers, datajsonlist, datadictlist
        

    @staticmethod
    def define_columns(dicdata, header1, header2):
        valueslist1 = []
        valueslist2 = []

        for data in dicdata:
            for k, v in data.items():
                if k == header1:
                    valueslist1.append(v)
                elif k == header2:
                    valueslist2.append(v)
        
        bnc_list, bn_list = ProcessData.modify_data(valueslist1, valueslist2)
        return bnc_list, bn_list


    @staticmethod
    def modify_data(bnc_data, bn_data):
        
        # obtener lista con los bnc's separados en caso de que existan mas de 2 en una sola casilla:
        bnc_clean = []
        for v in bnc_data:
            if '/' in v:
                splits_bnc = [e.strip() for e in v.split('/')]
                bnc_clean.extend(splits_bnc)
            else:
                bnc_clean.append(v)
        
        # obtener bn's separados en caso de que sean substring de un string:
        pattern = r'\bVA[A-Z0-9]{6}\b'
        bn_clean = [match.group() for item in bn_data for match in re.finditer(pattern, item)]

        return bnc_clean, bn_clean
    

    @staticmethod
    def cut_dict(dict_data, bnc_header, bn_header):
        new_data_dict = []
        pattern = r'\bVA[A-Z0-9]{6}\b'
        
        # En el siguiente bucle se esta definiendo una lista de diccionarios iterados, en donde la parte importante a recalcar
        # es que el 'bn' se ubicara adecuadamente en caso de que sean substring de un string:
        for dic in dict_data:
            bn_value = dic[bn_header]
            matched_bn = re.search(pattern, bn_value)  # Buscar coincidencia con el patrón

            new_dict = {
                dic[bnc_header]: {
                    'bnc': dic[bnc_header],
                    'bn': matched_bn.group() if matched_bn else None  # Asignar valor o None
                }
            }
            new_data_dict.append(new_dict)
        
        return new_data_dict


    
# pd = ProcessData("TransEsmeraldas_Airtable_Feb2025_t.xlsx", "TransEsmeraldas_Novo_Feb2025_t.xlsx")
# print(pd.cut_dictdata1, len(pd.cut_dictdata1))
# print(pd.cut_dictdata2)
